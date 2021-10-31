import os
import argparse
import json
import re
import pandas
from pandas.core.frame import DataFrame
from pandas.core.series import Series

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


def calculate(input_file: str, output_file: str, competition_type: str, break_tie: bool = False) -> bool:
    def find_feature_names(dataframe: DataFrame) -> dict[str: str]:
        feature_names = {}
        column_keys = dataframe.keys()

        # determine number column
        column_matches = column_keys[column_keys.str.match(r'.*number.*', case=False)]
        if len(column_matches) == 1:
            feature_names["number"] = column_matches[0]
        elif len(column_matches) > 1:
            raise KeyError("multiple 'number' columns detected")
        else:
            raise KeyError("no 'number' column detected")

        # determine handler column
        column_matches = column_keys[column_keys.str.match(r'.*handler.*', case=False)]
        if len(column_matches) == 1:
            feature_names["handler"] = column_matches[0]
        elif len(column_matches) > 1:
            raise KeyError("multiple 'handler' columns detected")
        else:
            raise KeyError("no 'handler' column detected")

        # determine call name column
        column_matches = column_keys[column_keys.str.match(r'.*call name.*', case=False)]
        if len(column_matches) == 1:
            feature_names["call name"] = column_matches[0]
        elif len(column_matches) > 1:
            raise KeyError("multiple 'call name' columns detected")
        else:
            raise KeyError("no 'call name' column detected")

        # determine class column
        column_matches = column_keys[column_keys.str.match(r'.*class.*', case=False)]
        if len(column_matches) == 1:
            feature_names["class"] = column_matches[0]
        elif len(column_matches) > 1:
            raise KeyError("multiple 'class' columns detected")
        else:
            raise KeyError("no 'class' column detected")

        # determine score column
        column_matches = column_keys[column_keys.str.match(r'.*score.*', case=False)]
        if len(column_matches) == 1:
            feature_names["score"] = column_matches[0]
        elif len(column_matches) > 1:
            raise KeyError("multiple 'score' columns detected")
        else:
            raise KeyError("no 'score' column detected")

        if competition_type == 'obedience':
            # determine champion column
            column_matches = column_keys[column_keys.str.match(r'.*champ.*', case=False)]
            if len(column_matches) == 1:
                feature_names["champion"] = column_matches[0]
            elif len(column_matches) > 1:
                raise KeyError("multiple 'champion' columns detected")

            # determine group column
            column_matches = column_keys[column_keys.str.match(r'.*group.*', case=False)]
            if len(column_matches) == 1:
                feature_names["group"] = column_matches[0]
            elif len(column_matches) > 1:
                raise KeyError("multiple 'group' columns detected")
            else:
                raise KeyError("no 'group' column detected")

        return feature_names

    def augment_data(contestants: DataFrame) -> DataFrame:
        """
        Removes nan rows and removes contestants with 'AB' or 'NQ' scores.
        :param contestants: a dataframe of contestants
        :return: a dataframe of contestants
        """

        # remove NaN entries and remove contestants who did not qualify or were absent
        contestants = contestants.loc[(~contestants[column_names["score"]].isna()) &
                                      (~contestants[column_names["score"]].str.match(r'\D+', na=False))].copy()

        # convert scores with + values to their base floats, with the count of +s in another column
        contestants.loc[:, 'Pluses'] = contestants[column_names["score"]].astype(str).str.count(r'\+')
        contestants.loc[:, column_names["score"]] = contestants[column_names["score"]].astype(str).replace(r'\++', '', regex=True).astype(float)

        rank_dict = {class_.lower(): len(settings["class hierarchy"]) - rank for rank, class_ in
                     enumerate(settings["class hierarchy"])}
        rank_dict[r".+"] = 0  # all other classes are ranked last
        contestants.loc[:, 'Class Rank'] = contestants[column_names["class"]].str.lower().replace(rank_dict, regex=True)

        return contestants

    def find_winners(contestants: DataFrame, condition: Series, is_award: bool, combine_scores=False,
                     break_tie: bool = False) -> DataFrame:

        # separate out contestants meeting condition and sort them based on numerical score
        contestants = contestants.loc[condition].sort_values(column_names["score"], ascending=False).copy()  # copy dataframe

        '''# if no contestants match criteria, return
        if contestants.shape[0] == 0:
            return contestants'''

        if is_award:
            # exclude beginner and preferred classes
            contestants = contestants.loc[
                (~contestants[column_names["class"]].str.contains(r'begin|preferred', case=False))].copy()

            # in case the contestants only had excluded classes
            if contestants.shape[0] == 0:
                return contestants

            # only used for specific awards
            if combine_scores:

                # pull out contestants that appear in all classes
                contestant_class_counts = contestants[column_names["call name"]].value_counts()
                contestants_in_all_classes = contestants[contestants[column_names["call name"]].isin(
                    contestant_class_counts.index[contestant_class_counts.ge(combine_scores)])]
                contestant_scores = contestants_in_all_classes.groupby(column_names["call name"])[column_names["score"]].sum().sort_values(
                    ascending=False)
                winning_contestants = contestant_scores.loc[contestant_scores == contestant_scores.max()]
                winning_info = contestants[contestants[column_names["call name"]].isin(winning_contestants.keys())]

                combined_info = winning_info[~winning_info[column_names["call name"]].duplicated()]
                combined_info.loc[:, column_names["score"]] = winning_contestants.max()

                if break_tie and combined_info.shape[0] > 1:
                    # sort based on class and scores
                    sorted_winning_contestants = contestants[
                        contestants[column_names["call name"]].isin(winning_contestants.keys())].sort_values(
                        ['Class Rank', column_names["score"], 'Pluses'], ascending=False)

                    return combined_info[combined_info[column_names["call name"]] == sorted_winning_contestants.iloc[0][column_names["call name"]]]

                else:
                    return combined_info

            else:
                # the highest scores amongst contestants
                contestants = contestants[contestants[column_names["score"]] == contestants[column_names["score"]].max()]
                if break_tie:  # break tie by class
                    score, class_rank, pluses = \
                        contestants.sort_values([column_names["score"], 'Class Rank', 'Pluses'], ascending=False).iloc[0][
                            [column_names["score"], 'Class Rank', 'Pluses']]
                    return contestants.loc[(contestants[column_names["score"]] == score) &
                                           (contestants['Class Rank'] == class_rank) &
                                           (contestants['Pluses'] == pluses)]
                else:
                    # if there is more than one class, don't break by pluses
                    if len(contestants[column_names["class"]].unique()) > 1:
                        return contestants
                    else:
                        score, pluses = contestants.sort_values([column_names["score"], 'Pluses'], ascending=False).iloc[0][
                            [column_names["score"], 'Pluses']]
                        return contestants.loc[(contestants[column_names["score"]] == score) &
                                               (contestants['Pluses'] == pluses)]

        # all the same class, return top 4, tie breaking done by pluses
        else:
            return contestants.sort_values([column_names["score"], 'Pluses'], ascending=False)[:4]

    def write_winners(workbook: Workbook, contestants: DataFrame, competition_name: str, is_award: bool) -> None:
        if is_award:
            if contestants.shape[0] < 1:
                workbook.append([competition_name] + ['-'] * 3)
                workbook[f"A{workbook.max_row}"].font = Font(bold=True)

            else:
                for c, contestant in enumerate(
                        dataframe_to_rows(contestants[[column_names["number"], column_names["handler"], column_names["call name"]]], index=False,
                                          header=False)):
                    if c == 0:
                        workbook.append([competition_name] + contestant)

                        # format competition name
                        workbook[f"A{workbook.max_row}"].font = Font(bold=True)
                    else:
                        workbook.append([''] + contestant)

        else:
            # write competition name and format cells
            workbook.append(['', competition_name])
            workbook[f"B{workbook.max_row}"].font = Font(bold=True)
            workbook[f"B{workbook.max_row}"].alignment = Alignment(horizontal='center')
            workbook.merge_cells(start_row=workbook.max_row, start_column=2, end_row=workbook.max_row, end_column=4)

            # write subheaders
            workbook.append(['', column_names["number"], column_names["handler"], column_names["call name"]])
            workbook[f"B{workbook.max_row}"].font = Font(bold=True)  # TODO: probably a better way to do more than one
            workbook[f"C{workbook.max_row}"].font = Font(bold=True)
            workbook[f"D{workbook.max_row}"].font = Font(bold=True)

            color_key = {0: '000000FF', 1: '00FF0000', 2: '00FFFF00', 3: '00FFFFFF'}
            text_key = {0: '00FFFFFF', 1: '00FFFFFF', 2: '00000000', 3: '00000000'}

            # write winners
            if contestants.shape[0] < 1:
                workbook.append([''] + ['-'] * 3)
            else:
                position = list(range(1, contestants.shape[0] + 1))
                for c, contestant in enumerate(
                        dataframe_to_rows(contestants[[column_names["number"], column_names["handler"], column_names["call name"]]], index=False,
                                          header=False)):
                    workbook.append([position[c]] + contestant)

                    workbook[f"B{workbook.max_row}"].font = Font(color=text_key[c])
                    workbook[f"B{workbook.max_row}"].fill = PatternFill(start_color=color_key[c],
                                                                        end_color=color_key[c], fill_type='solid')
                    workbook[f"C{workbook.max_row}"].font = Font(color=text_key[c])
                    workbook[f"C{workbook.max_row}"].fill = PatternFill(start_color=color_key[c],
                                                                        end_color=color_key[c], fill_type='solid')
                    workbook[f"D{workbook.max_row}"].font = Font(color=text_key[c])
                    workbook[f"D{workbook.max_row}"].fill = PatternFill(start_color=color_key[c],
                                                                        end_color=color_key[c], fill_type='solid')

            workbook.append([''])

    def optimal_width(sheet: Worksheet) -> None:
        """
        Formats the worksheet so each column is the width of the largest cell value.
        :param sheet: the worksheet to resize
        """
        for column in sheet.columns:
            new_column_length = max(len(str(cell.value)) for cell in column)  # find max cell width in column
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            sheet.column_dimensions[column_letter].width = new_column_length

    # make sure input file is valid
    if input_file[-5:] != '.xlsx':
        raise TypeError("only .xlsx is supported")

    # check to see if file exists
    if not os.path.exists(input_file):
        raise FileNotFoundError("file not found")

    # import settings
    if os.path.exists("settings.json"):
        with open("settings.json", 'r') as f:
            settings = json.load(f)
    else:
        with open("settings.json", 'w') as f:
            json.dump({"class hierarchy": ["utility", "open", "novice"],
                       "defaults": {"break ties by class": False, "write to new file": False}}, f)

        settings = json.load(f)

    data = pandas.read_excel(input_file)

    # find column names
    column_names = find_feature_names(data)

    classes = data[column_names["class"]].dropna().unique()  # find set of classes

    data = augment_data(data)  # remove nan rows and contestants with non-scores

    if input_file == output_file:
        output_workbook = openpyxl.load_workbook(input_file)
        output_workbook.create_sheet("Winners")
        output_worksheet = output_workbook["Winners"]
    else:
        output_workbook = Workbook()
        output_worksheet = output_workbook.active

    for class_ in classes:
        winners = find_winners(data, Series(data[column_names["class"]] == class_), False, break_tie=break_tie)
        write_winners(output_worksheet, winners, class_, False)

    # add headers for award winners
    output_worksheet.append(['', column_names["number"], column_names["handler"], column_names["call name"]])
    output_worksheet[f'B{output_worksheet.max_row}'].font = Font(bold=True)
    output_worksheet[f'C{output_worksheet.max_row}'].font = Font(bold=True)
    output_worksheet[f'D{output_worksheet.max_row}'].font = Font(bold=True)

    if competition_type == "obedience":
        award_conditions = [
            ("High in Trial", data[column_names["class"]].str.contains(r"\sb$", case=False), False),
            ("High Combined", ((data[column_names["class"]].str.contains("open b", case=False)) |
                               (data[column_names["class"]].str.contains("utility b", case=False))), 2),
            ("High Combined Preferred",
             ((data[column_names["class"]].str.contains("preferred open", case=False)) |
              (data[column_names["class"]].str.contains("preferred utility", case=False))), 2)
        ]
        if column_names["champion"] is not None:
            award_conditions.append(
                ("High Scoring Champion of Record", data[column_names["champion"]].str.contains('Ch', na=False), False)
            )

    elif competition_type == "rally":
        award_conditions = (
            ("High Combined", ((data[column_names["class"]].str.contains(r"r(ally)? excellent b", case=False)) |
                               (data[column_names["class"]].str.contains(r"r(ally)? advanced? b", case=False))), 2),
            ("High Triple", ((data[column_names["class"]].str.contains(r"r(ally)? excellent b", case=False)) |
                             (data[column_names["class"]].str.contains(r"r(ally)? advanced? b", case=False)) |
                             (data[column_names["class"]].str.contains(r"r(ally)? master", case=False))), 3)
        )
    else:
        raise ValueError("competition_type must be 'obedience' or 'rally'")

    # find award winners
    for award, award_condition, combine_scores in award_conditions:
        winners = find_winners(data, award_condition, True, combine_scores=combine_scores, break_tie=break_tie)
        write_winners(output_worksheet, winners, award, True)

    if competition_type == "obedience":
        groups = [group for group in data[column_names["group"]].unique() if str(group) != 'nan']  # find set of groups

        # find group winners
        for group in groups:
            winners = find_winners(data, Series(data[column_names["group"]] == group), True, break_tie=break_tie)
            write_winners(output_worksheet, winners, group, True)

    optimal_width(output_worksheet)

    output_workbook.save(output_file)

    return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Determine the winners of a dog show.")
    parser.add_argument('input_file', type=open, help="the input .xlsx file")
    parser.add_argument('output_file', help="the output file")
    parser.add_argument('competition_type', choices=['obedience', 'rally'], help="the type of competition")
    parser.add_argument('-n', '--no_ties', action='store_true', help="break ties by class if possible")
    args = parser.parse_args()

    status = calculate(args.input_file.name, args.output_file, args.competition_type, args.no_ties)

    if status:
        print("Complete")
    else:
        print(f"Unknown Error")
